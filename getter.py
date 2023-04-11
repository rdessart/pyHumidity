import sys
import sqlalchemy
import requests

from sqlalchemy.orm import declarative_base
from datetime import datetime
from openpyxl import Workbook
from statistics import mean

engine = sqlalchemy.create_engine("mariadb+mariadbconnector://rdessart:roro3323@192.168.0.10:3306/home")
Base = sqlalchemy.orm.declarative_base()
class HumidityValue(Base):
    __tablename__ = 'humidity'
    id = sqlalchemy.Column(sqlalchemy.Text, primary_key=True)
    datetime = sqlalchemy.Column(sqlalchemy.DateTime)
    temperature = sqlalchemy.Column(sqlalchemy.Double)
    humidity = sqlalchemy.Column(sqlalchemy.Double)

    def Serialize(self):
        return {
            "Timestamp" : self.datetime.isoformat(),
            "Temperature": self.temperature,
            "Humidity": self.humidity,
        }


class HumidityValue2:
    def __init__(self, data):
        self.EntryId = 0
        self.TimeStamp = datetime.utcnow()
        self.Temperature = 0.0
        self.Humidity  = 0.0
        if data is not None:
            self.EntryId = data["entryId"]
            self.TimeStamp = datetime.strptime(data["timeStamp"], "%Y-%m-%dT%H:%M:%S.%fZ")
            self.Temperature = data["temperature"]
            self.Humidity = data["humidity"]
        

    def Serialize(self):
        return {
            "Timestamp" : self.datetime.isoformat(),
            "Temperature": self.temperature,
            "Humidity": self.humidity,
        }

def GetAndParseData1(book, averages):
    Session = sqlalchemy.orm.sessionmaker()
    Session.configure(bind=engine)
    session = Session()

    datas: list[HumidityValue] = session.query(HumidityValue).all()
    datas_dict: dict[datetime, list[HumidityValue]] = {}


    #arranging per date
    for data in datas:
        date = data.datetime.date()
        if date not in datas_dict:
            datas_dict[date] = []
        datas_dict[date].append(data)

    for date in list(datas_dict.keys()):
        print(f"Parsing {date}")
        sheet = book.create_sheet(date.isoformat())
        sheet["A1"] = "TIME"
        sheet["B1"] = "HUMIDITY"
        sheet["C1"] = "TEMPERATURE"
        average_h = []
        average_t = []
        for row, data in enumerate(list(datas_dict[date])):
            sheet.cell(row=row+2, column=1, value=data.datetime)
            sheet.cell(row=row+2, column=2, value=data.humidity)
            sheet.cell(row=row+2, column=3, value=data.temperature)
            average_h.append(data.humidity)
            average_t.append(data.temperature)

        averages[date] = HumidityValue2(None)
        averages[date].Humidity = mean(average_h)
        averages[date].Temperature = mean(average_t)


def GetAndParseData2(book, averages):
    url = "https://humidity.e-skystudio.com/api/Humidity"
    r = requests.get(url=url)
    # print(HumidityValue(r.json()[0]))
    datas = [HumidityValue2(d) for d in r.json()]
    datas_dict: dict[datetime, list[HumidityValue]] = {}


    #arranging per date
    for data in datas:
        date = data.TimeStamp.date()
        if date not in datas_dict:
            datas_dict[date] = []
        datas_dict[date].append(data)

    for date in list(datas_dict.keys()):
        print(f"Parsing {date}")
        sheet = book.create_sheet(date.isoformat())
        sheet["A1"] = "TIME"
        sheet["B1"] = "HUMIDITY"
        sheet["C1"] = "TEMPERATURE"
        average_h = []
        average_t = []
        for row, data in enumerate(list(datas_dict[date])):
            sheet.cell(row=row+2, column=1, value=data.TimeStamp)
            sheet.cell(row=row+2, column=2, value=data.Humidity)
            sheet.cell(row=row+2, column=3, value=data.Temperature)
            average_h.append(data.Humidity)
            average_t.append(data.Temperature)

        averages[date] = HumidityValue2(None)
        averages[date].Humidity = mean(average_h)
        averages[date].Temperature = mean(average_t)

def average(book, average):
    sheet = book.create_sheet("AVERAGE", 0)
    sheet.cell(row=1, column=1, value="Date")
    sheet.cell(row=1, column=2, value="Humidity")
    sheet.cell(row=1, column=3, value="Temperature")
    for id, [date, data] in enumerate(averages.items()):
        sheet.cell(row=id+2, column=1, value=date)
        sheet.cell(row=id+2, column=2, value=data.Humidity)
        sheet.cell(row=id+2, column=3, value=data.Temperature)


if __name__ == "__main__":
    book = Workbook()
    book.remove(book.active)
    averages :dict[datetime, HumidityValue2] = {}
    print("Parsing DB 1...")
    GetAndParseData1(book, averages)
    print("Parsing DB 2...")
    GetAndParseData2(book, averages)
    print("Adding averages...")
    average(book, averages)

    book.save("humidity3.xlsx")